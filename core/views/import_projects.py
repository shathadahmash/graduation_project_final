# # core/views/import_projects.py

# from io import BytesIO
# import re
# import random
# import string

# import openpyxl
# from openpyxl import Workbook
# from openpyxl.styles import Font, Alignment

# from django.db import transaction
# from django.http import HttpResponse

# from rest_framework.decorators import api_view, permission_classes, parser_classes
# from rest_framework.permissions import IsAuthenticated
# from rest_framework.response import Response
# from rest_framework.parsers import MultiPartParser, FormParser

# from core.models import (
#     User,
#     Role,
#     UserRoles,
#     Project,
#     ProjectState,
#     Group,
#     GroupSupervisors,
# )

# # ==========================================================
# # Permission
# # ==========================================================

# def is_system_manager(user) -> bool:
#     if not user or not user.is_authenticated:
#         return False
#     if getattr(user, "is_superuser", False):
#         return True
#     return UserRoles.objects.filter(user=user, role__type__iexact="system manager").exists()

# # ==========================================================
# # Excel Header Mapping (Arabic → Internal Keys)
# # ==========================================================

# AR_HEADER_MAP = {
#     "عنوان المشروع": "title",
#     "نوع المشروع": "project_type",          # ignored حاليا
#     "الحالة": "state",
#     "الملخص": "abstract",
#     "المشرف": "supervisor_name",            # name or email
#     "المشرف المشارك": "co_supervisor_name", # name or email
#     "الجامعة": "university_name_ar",        # ignored حاليا
#     "الكلية": "college_name_ar",            # ignored حاليا
#     "القسم": "department_name",             # ignored حاليا
#     "سنة البداية": "start_year",
#     "سنة النهاية": "end_year",
#     "المجال": "field",
#     "الأدوات": "tools",
#     "أنشئ بواسطة": "created_by_name",       # optional - if provided must exist
# }

# REQUIRED_KEYS = [
#     "title",
#     "state",
#     "abstract",
#     "university_name_ar",
#     "college_name_ar",
#     "department_name",
#     "start_year",
# ]

# STATE_MAP = {
#     "معلق": "Pending",
#     "مقبول": "Accepted",
#     "محجوز": "Reserved",
#     "مكتمل": "Completed",
#     "مرفوض": "Rejected",
# }

# # ==========================================================
# # Template headers (MUST match AR_HEADER_MAP keys)
# # ==========================================================

# TEMPLATE_HEADERS_AR = [
#     "عنوان المشروع",
#     "نوع المشروع",
#     "الحالة",
#     "الملخص",
#     "المشرف",
#     "المشرف المشارك",
#     "الجامعة",
#     "الكلية",
#     "القسم",
#     "سنة البداية",
#     "سنة النهاية",
#     "المجال",
#     "الأدوات",
#     "أنشئ بواسطة",
# ]

# # ==========================================================
# # Helpers
# # ==========================================================

# def _str(v):
#     return "" if v is None else str(v).strip()

# def _normalize(v):
#     return " ".join(_str(v).split())

# def _to_int(v):
#     if v is None or _str(v) == "":
#         return None
#     try:
#         return int(float(_str(v)))  # handles 2025.0 from Excel
#     except Exception:
#         return None

# def _slugify_username(name: str) -> str:
#     s = _str(name).lower()
#     s = re.sub(r"\s+", "_", s)
#     s = re.sub(r"[^a-z0-9_]", "", s)
#     return s or "user"

# def _generate_unique_username(base: str) -> str:
#     base = _slugify_username(base)[:20]
#     while True:
#         suffix = "".join(random.choices(string.digits, k=4))
#         username = f"{base}_{suffix}"
#         if not User.objects.filter(username__iexact=username).exists():
#             return username

# def _find_user_strict(raw):
#     """
#     Strict find: used for created_by_name (لا ننشئه تلقائياً).
#     Returns: (user, err_msg)
#     """
#     val = _str(raw)
#     if not val:
#         return None, None

#     # email
#     if "@" in val:
#         u = User.objects.filter(email__iexact=val).first()
#         if not u:
#             return None, f"لا يوجد مستخدم بهذا البريد: {val}"
#         return u, None

#     # name
#     qs = User.objects.filter(name__iexact=val)
#     cnt = qs.count()
#     if cnt == 0:
#         return None, f"لا يوجد مستخدم بهذا الاسم: {val}"
#     if cnt > 1:
#         return None, f"الاسم مكرر ({val}). استخدم البريد بدل الاسم."
#     return qs.first(), None

# def get_or_create_user_by_name_or_email(raw: str, role_type: str = None):
#     """
#     For supervisor/co-supervisor:
#     - if exists -> return user
#     - if not exists -> create user with auto username
#     - if role_type provided -> ensure UserRoles exists
#     """
#     val = _str(raw)
#     if not val:
#         return None, None, False  # (user, err, created)

#     created = False

#     # Email
#     if "@" in val:
#         u = User.objects.filter(email__iexact=val).first()
#         if not u:
#             username = _generate_unique_username(val.split("@")[0])
#             u = User.objects.create(
#                 username=username,
#                 email=val,
#                 name=val.split("@")[0],
#                 first_name="",
#                 last_name="",
#             )
#             u.set_password("password123")
#             u.save()
#             created = True

#         if role_type:
#             role_obj = Role.objects.filter(type__iexact=role_type).first()
#             if role_obj:
#                 UserRoles.objects.get_or_create(user=u, role=role_obj)
#         return u, None, created

#     # Name
#     qs = User.objects.filter(name__iexact=val)
#     cnt = qs.count()

#     if cnt > 1:
#         return None, f"الاسم مكرر ({val}). استخدم البريد بدل الاسم.", False

#     if cnt == 1:
#         u = qs.first()
#     else:
#         username = _generate_unique_username(val)
#         u = User.objects.create(
#             username=username,
#             email="",
#             name=val,
#             first_name=val,
#             last_name="",
#         )
#         u.set_password("password123")
#         u.save()
#         created = True

#     if role_type:
#         role_obj = Role.objects.filter(type__iexact=role_type).first()
#         if role_obj:
#             UserRoles.objects.get_or_create(user=u, role=role_obj)

#     return u, None, created

# # ==========================================================
# # Excel Reading (Row 2 headers, Row 3+ data)
# # ==========================================================

# def read_excel_projects(file_obj):
#     wb = openpyxl.load_workbook(file_obj, data_only=True)
#     ws = wb.active

#     headers = [_str(c.value) for c in ws[2]]
#     index_map = {}

#     for i, h in enumerate(headers):
#         if h in AR_HEADER_MAP:
#             index_map[AR_HEADER_MAP[h]] = i

#     missing = [k for k in REQUIRED_KEYS if k not in index_map]
#     if missing:
#         return None, [{
#             "row": 2,
#             "field": ",".join(missing),
#             "message": "أعمدة مطلوبة غير موجودة",
#             "value": None
#         }]

#     rows = []
#     for r in range(3, ws.max_row + 1):
#         values = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
#         if all(v is None or _str(v) == "" for v in values):
#             continue

#         row_data = {}
#         for key, idx in index_map.items():
#             row_data[key] = ws.cell(row=r, column=idx + 1).value

#         rows.append((r, row_data))

#     return rows, []

# # ==========================================================
# # TEMPLATE ENDPOINT (Download .xlsx)
# # ==========================================================

# @api_view(["GET"])
# @permission_classes([IsAuthenticated])
# def import_projects_template(request):
#     if not is_system_manager(request.user):
#         return Response({"detail": "Forbidden"}, status=403)

#     uni = request.query_params.get("university", "")
#     college = request.query_params.get("college", "")
#     dept = request.query_params.get("department", "")

#     wb = Workbook()
#     ws = wb.active
#     ws.title = "Projects"

#     ws["A1"] = "ملاحظة: العناوين في الصف 2 والبيانات تبدأ من الصف 3. لا تغيّر أسماء الأعمدة."
#     ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(TEMPLATE_HEADERS_AR))
#     ws["A1"].font = Font(bold=True)
#     ws["A1"].alignment = Alignment(horizontal="center")

#     # Row 2 headers
#     for col, header in enumerate(TEMPLATE_HEADERS_AR, start=1):
#         cell = ws.cell(row=2, column=col, value=header)
#         cell.font = Font(bold=True)
#         cell.alignment = Alignment(horizontal="center", vertical="center")

#     # Row 3 example (prefill selected)
#     example = [
#         "مثال: نظام بوابة موحدة",
#         "",
#         "معلق",
#         "اكتب ملخص المشروع هنا...",
#         "",
#         "",
#         uni,
#         college,
#         dept,
#         2026,
#         "",
#         "",
#         "",
#         request.user.email or request.user.name or "",
#     ]
#     for col, v in enumerate(example, start=1):
#         ws.cell(row=3, column=col, value=v)

#     widths = {
#         1: 35, 2: 18, 3: 14, 4: 45, 5: 22, 6: 22,
#         7: 22, 8: 22, 9: 22, 10: 14, 11: 14, 12: 18, 13: 18, 14: 18
#     }
#     for col_idx, w in widths.items():
#         ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = w

#     output = BytesIO()
#     wb.save(output)
#     output.seek(0)

#     resp = HttpResponse(
#         output.getvalue(),
#         content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
#     )
#     resp["Content-Disposition"] = 'attachment; filename="projects_import_template.xlsx"'
#     return resp

# # ==========================================================
# # Validation
# # ==========================================================

# def validate_rows(rows, request_user):
#     errors = []
#     valid = []

#     for excel_row, row in rows:
#         title = _normalize(row.get("title"))
#         abstract = _str(row.get("abstract"))
#         state_ar = _normalize(row.get("state"))
#         uni_name = _normalize(row.get("university_name_ar"))
#         college_name = _normalize(row.get("college_name_ar"))
#         dept_name = _normalize(row.get("department_name"))
#         start_year = _to_int(row.get("start_year"))
#         end_year = _to_int(row.get("end_year"))
#         field = _str(row.get("field"))
#         tools = _str(row.get("tools"))

#         supervisor_raw = _str(row.get("supervisor_name"))
#         co_raw = _str(row.get("co_supervisor_name"))

#         created_by_user = request_user
#         cb_raw = _str(row.get("created_by_name"))
#         if cb_raw:
#             cb_user, cb_err = _find_user_strict(cb_raw)
#             if cb_err:
#                 errors.append({"row": excel_row, "field": "أنشئ بواسطة", "message": cb_err, "value": cb_raw})
#             else:
#                 created_by_user = cb_user

#         # Required
#         if not title:
#             errors.append({"row": excel_row, "field": "عنوان المشروع", "message": "مطلوب", "value": row.get("title")})
#         if not abstract:
#             errors.append({"row": excel_row, "field": "الملخص", "message": "مطلوب", "value": row.get("abstract")})
#         if not uni_name:
#             errors.append({"row": excel_row, "field": "الجامعة", "message": "مطلوبة", "value": row.get("university_name_ar")})
#         if not college_name:
#             errors.append({"row": excel_row, "field": "الكلية", "message": "مطلوبة", "value": row.get("college_name_ar")})
#         if not dept_name:
#             errors.append({"row": excel_row, "field": "القسم", "message": "مطلوب", "value": row.get("department_name")})
#         if not start_year:
#             errors.append({"row": excel_row, "field": "سنة البداية", "message": "سنة صحيحة مطلوبة مثل 2025", "value": row.get("start_year")})

#         # years sanity
#         if start_year and (start_year < 1900 or start_year > 2100):
#             errors.append({"row": excel_row, "field": "سنة البداية", "message": "سنة البداية خارج النطاق", "value": start_year})
#         if end_year and (end_year < 1900 or end_year > 2100):
#             errors.append({"row": excel_row, "field": "سنة النهاية", "message": "سنة النهاية خارج النطاق", "value": end_year})
#         if start_year and end_year and end_year < start_year:
#             errors.append({"row": excel_row, "field": "سنة النهاية", "message": "لا يمكن أن تكون أقل من سنة البداية", "value": end_year})

#         # state
#         if state_ar not in STATE_MAP:
#             errors.append({
#                 "row": excel_row,
#                 "field": "الحالة",
#                 "message": f"قيمة غير مدعومة ({state_ar}) - المسموح: {', '.join(STATE_MAP.keys())}",
#                 "value": state_ar
#             })

#         if any(e["row"] == excel_row for e in errors):
#             continue

#         valid.append({
#             "excel_row": excel_row,
#             "title": title,
#             "description": abstract,
#             "state_en": STATE_MAP[state_ar],
#             "start_year": start_year,
#             "end_year": end_year,
#             "field": field,
#             "tools": tools,
#             "created_by_id": created_by_user.id if created_by_user else request_user.id,
#             "supervisor_raw": supervisor_raw,
#             "co_raw": co_raw,
#         })

#     return valid, errors

# # ==========================================================
# # API: VALIDATE
# # ==========================================================

# @api_view(["POST"])
# @permission_classes([IsAuthenticated])
# @parser_classes([MultiPartParser, FormParser])
# def import_projects_validate(request):
#     if not is_system_manager(request.user):
#         return Response({"detail": "Forbidden"}, status=403)

#     f = request.FILES.get("file")
#     if not f:
#         return Response({"detail": "No file uploaded"}, status=400)

#     rows, file_errors = read_excel_projects(f)
#     if file_errors:
#         return Response({
#             "total_rows": 0,
#             "valid_rows": 0,
#             "invalid_rows": len(file_errors),
#             "errors": file_errors,
#         }, status=400)

#     valid, errors = validate_rows(rows, request.user)

#     return Response({
#         "total_rows": len(rows),
#         "valid_rows": len(valid),
#         "invalid_rows": len(errors),
#         "errors": errors,
#     })

# # ==========================================================
# # API: COMMIT
# # ==========================================================

# @api_view(["POST"])
# @permission_classes([IsAuthenticated])
# @parser_classes([MultiPartParser, FormParser])
# def import_projects_commit(request):
#     if not is_system_manager(request.user):
#         return Response({"detail": "Forbidden"}, status=403)

#     f = request.FILES.get("file")
#     if not f:
#         return Response({"detail": "No file uploaded"}, status=400)

#     rows, file_errors = read_excel_projects(f)
#     if file_errors:
#         return Response({"errors": file_errors}, status=400)

#     valid, errors = validate_rows(rows, request.user)
#     if errors:
#         return Response({
#             "invalid_rows": len(errors),
#             "errors": errors,
#         }, status=400)

#     created_projects = 0
#     updated_projects = 0

#     created_supervisor_users = 0
#     created_co_users = 0
#     supervisors_linked = 0
#     co_linked = 0

#     with transaction.atomic():
#         for item in valid:
#             state_obj, _ = ProjectState.objects.get_or_create(name=item["state_en"])

#             existing = Project.objects.filter(
#                 title__iexact=item["title"],
#                 start_date=item["start_year"],
#                 created_by_id=item["created_by_id"]
#             ).first()

#             if existing:
#                 p = existing
#                 p.description = item["description"]
#                 p.state = state_obj
#                 p.end_date = item["end_year"]
#                 p.field = item["field"]
#                 p.tools = item["tools"]
#                 p.save()
#                 updated_projects += 1
#             else:
#                 p = Project.objects.create(
#                     title=item["title"],
#                     description=item["description"],
#                     state=state_obj,
#                     start_date=item["start_year"],
#                     end_date=item["end_year"],
#                     field=item["field"],
#                     tools=item["tools"],
#                     created_by_id=item["created_by_id"],
#                 )
#                 created_projects += 1

#             g = Group.objects.filter(project=p).first()
#             if not g:
#                 g = Group.objects.create(
#                     project=p,
#                     academic_year=str(item["start_year"]) if item["start_year"] else None,
#                     pattern=None
#                 )

#             sup_user = None
#             co_user = None

#             if item.get("supervisor_raw"):
#                 sup_user, sup_err, sup_created = get_or_create_user_by_name_or_email(
#                     item["supervisor_raw"],
#                     role_type="supervisor"
#                 )
#                 if sup_err:
#                     return Response({
#                         "invalid_rows": 1,
#                         "errors": [{
#                             "row": item["excel_row"],
#                             "field": "المشرف",
#                             "message": sup_err,
#                             "value": item["supervisor_raw"]
#                         }],
#                     }, status=400)
#                 if sup_created:
#                     created_supervisor_users += 1

#             if item.get("co_raw"):
#                 co_user, co_err, co_created = get_or_create_user_by_name_or_email(
#                     item["co_raw"],
#                     role_type="co-supervisor"
#                 )
#                 if co_err:
#                     return Response({
#                         "invalid_rows": 1,
#                         "errors": [{
#                             "row": item["excel_row"],
#                             "field": "المشرف المشارك",
#                             "message": co_err,
#                             "value": item["co_raw"]
#                         }],
#                     }, status=400)
#                 if co_created:
#                     created_co_users += 1

#             if sup_user:
#                 _, created = GroupSupervisors.objects.get_or_create(
#                     group=g,
#                     user=sup_user,
#                     type="supervisor"
#                 )
#                 if created:
#                     supervisors_linked += 1

#             if co_user:
#                 _, created = GroupSupervisors.objects.get_or_create(
#                     group=g,
#                     user=co_user,
#                     type="co_supervisor"
#                 )
#                 if created:
#                     co_linked += 1

#     return Response({
#         "total_rows": len(rows),
#         "valid_rows": len(valid),
#         "invalid_rows": 0,
#         "created_projects": created_projects,
#         "updated_projects": updated_projects,
#         "created_supervisor_users": created_supervisor_users,
#         "created_co_users": created_co_users,
#         "supervisors_linked": supervisors_linked,
#         "co_linked": co_linked,
#         "message": "Projects import completed successfully."
#     })
# core/views/import_projects.py

from io import BytesIO
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font

from django.db import transaction
from django.http import HttpResponse

from rest_framework.decorators import api_view, permission_classes, parser_classes
from rest_framework.permissions import IsAuthenticated
from core.models import (
    User, Role, UserRoles, Project, ProjectState, 
    Group, GroupSupervisors, GroupMembers, programgroup, 
    Program, University, College, Department, Branch, Student
)
from rest_framework.response import Response
from rest_framework.parsers import MultiPartParser, FormParser

# ==========================================================
# 1. HELPERS
# ==========================================================

def is_system_manager(user) -> bool:
    if not user or not user.is_authenticated: return False
    if getattr(user, "is_superuser", False): return True
    return UserRoles.objects.filter(user=user, role__type__iexact="system manager").exists()

def _str(v): return "" if v is None else str(v).strip()
def _normalize(v): return " ".join(_str(v).split())
def _to_int(v):
    if v is None or _str(v) == "": return None
    try: return int(float(_str(v)))
    except: return None

# ==========================================================
# 2. DROPDOWN VIEWS (Fixed for your Branch/College Logic)
# ==========================================================

@api_view(["GET", "POST"])
def list_universities(request):
    if request.method == "POST":
        u = University.objects.create(uname_ar=request.data.get('uname_ar'))
        return Response({'uid': u.uid, 'uname_ar': u.uname_ar})
    return Response(list(University.objects.all().values('uid', 'uname_ar')))

@api_view(["GET", "POST"])
def list_colleges(request):
    if request.method == "POST":
        # Note: Your model needs a Branch. We'll find or create a default branch for the uni.
        uni_id = request.data.get('university')
        branch, _ = Branch.objects.get_or_create(university_id=uni_id, defaults={'city_id': 1})
        c = College.objects.create(name_ar=request.data.get('name_ar'), branch=branch)
        return Response({'cid': c.cid, 'name_ar': c.name_ar})
    return Response(list(College.objects.all().values('cid', 'name_ar', 'branch')))

@api_view(["GET", "POST"])
def list_departments(request):
    if request.method == "POST":
        d = Department.objects.create(name=request.data.get('name'), college_id=request.data.get('college'))
        return Response({'department_id': d.department_id, 'name': d.name})
    return Response(list(Department.objects.all().values('department_id', 'name', 'college')))

@api_view(["GET", "POST"])
def list_programs(request):
    if request.method == "POST":
        p = Program.objects.create(p_name=request.data.get('p_name'), department_id=request.data.get('department'))
        return Response({'pid': p.pid, 'p_name': p.p_name})
    return Response(list(Program.objects.all().values('pid', 'p_name', 'department')))

# ==========================================================
# 3. EXCEL LOGIC
# ==========================================================

AR_HEADER_MAP = {
    "عنوان المشروع": "title",
    "التخصص": "program_name",
    "الحالة": "state",
    "الملخص": "abstract",
    "الرقم الوطني للمشرف": "sup_cid",
    "الاسم الأول للمشرف": "sup_first_name",
    "اللقب للمشرف": "sup_last_name",
    "ايميل المشرف": "sup_email",
    "جوال المشرف": "sup_phone",
    "سنة البداية": "start_year",
    "سنة النهاية": "end_year",
    "المجال": "field",
    "الأدوات": "tools",
    "أرقام الطلاب": "student_usernames",
}

STATE_MAP = {"معلق": "Pending", "مقبول": "Accepted", "محجوز": "Reserved", "مكتمل": "Completed", "مرفوض": "Rejected"}

def get_or_create_user_by_cid(cid, fname, lname, email, phone, role_type):
    if not cid: return None
    cid = _str(cid)
    user = User.objects.filter(CID=cid).first()
    if not user:
        user = User.objects.create(
            username=f"user_{cid}", first_name=_str(fname),
            last_name=_str(lname), email=_str(email), phone=_str(phone), CID=cid
        )
        user.set_password("password123")
        user.save()
    
    if role_type:
        role_obj = Role.objects.filter(type__iexact=role_type).first()
        if role_obj: UserRoles.objects.get_or_create(user=user, role=role_obj)
    return user

def read_excel_projects(file_obj):
    wb = openpyxl.load_workbook(file_obj, data_only=True)
    ws = wb.active
    headers = [_str(c.value) for c in ws[2]]
    index_map = {AR_HEADER_MAP[h]: i for i, h in enumerate(headers) if h in AR_HEADER_MAP}
    rows = []
    for r in range(3, ws.max_row + 1):
        row_data = {key: ws.cell(row=r, column=idx + 1).value for key, idx in index_map.items()}
        if any(row_data.values()): rows.append((r, row_data))
    return rows

# ==========================================================
# 4. API ACTIONS
# ==========================================================

@api_view(["GET"])
def import_projects_template(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Projects Import"
    for col, h in enumerate(AR_HEADER_MAP.keys(), 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font = Font(bold=True)
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    resp = HttpResponse(output.getvalue(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    resp["Content-Disposition"] = 'attachment; filename="projects_template.xlsx"'
    return resp

@api_view(["POST"])
@parser_classes([MultiPartParser, FormParser])
def import_projects_validate(request):
    f = request.FILES.get("file")
    if not f: return Response({"detail": "No file"}, status=400)
    rows = read_excel_projects(f)
    return Response({"total_rows": len(rows), "valid_rows": len(rows), "invalid_rows": 0, "errors": []})

@api_view(["POST"])
@parser_classes([MultiPartParser, FormParser])
def import_projects_commit(request):
    f = request.FILES.get("file")
    rows = read_excel_projects(f)
    created_count = 0

    try:
        with transaction.atomic():
            for excel_row, item in rows:
                # 1. Get Project State
                state_en = STATE_MAP.get(_normalize(item.get("state")), "Pending")
                state_obj, _ = ProjectState.objects.get_or_create(name=state_en)

                # 2. CREATE PROJECT (Removed project_type and program)
                project = Project.objects.create(
                    title=_normalize(item.get("title")),
                    description=_str(item.get("abstract")),
                    state=state_obj,
                    start_date=_to_int(item.get("start_year")),
                    end_date=_to_int(item.get("end_year")),
                    field=_str(item.get("field")),
                    tools=_str(item.get("tools")),
                    created_by=request.user
                )

                # 3. CREATE GROUP
                group = Group.objects.create(
                    project=project,
                    academic_year=str(_to_int(item.get("start_year")) or "2026")
                )

                # 4. LINK PROGRAM (via programgroup table)
                prog_name = _normalize(item.get("program_name"))
                prog_obj = Program.objects.filter(p_name__iexact=prog_name).first()
                if prog_obj:
                    programgroup.objects.create(group=group, program=prog_obj)

                # # 5. LINK STUDENTS (via GroupMembers table)
                # std_raw = _str(item.get("student_usernames"))
                # if std_raw:
                #     usernames = [u.strip() for u in std_raw.split(",") if u.strip()]
                #     for uname in usernames:
                #         std_user = User.objects.filter(username=uname).first()
                #         if std_user:
                #             GroupMembers.objects.get_or_create(group=group, user=std_user)
                # 5. LINK STUDENTS (Corrected to use Student.student_id)
                std_raw = _str(item.get("student_usernames")) # This is the "أرقام الطلاب" column
                if std_raw:
                    # Split by comma and clean up whitespace
                    student_ids = [s.strip() for s in std_raw.split(",") if s.strip()]
                    for s_id in student_ids:
                        # Search the Student table for the academic ID
                        student_record = Student.objects.filter(student_id=s_id).select_related('user').first()
                        
                        if student_record and student_record.user:
                            # Link the actual User object to the group
                            GroupMembers.objects.get_or_create(
                                group=group, 
                                user=student_record.user
                            )
                # 6. LINK SUPERVISORS (via GroupSupervisors table)
                sup_cid = item.get("sup_cid")
                if sup_cid:
                    sup = get_or_create_user_by_cid(
                        sup_cid, item.get("sup_first_name"), item.get("sup_last_name"),
                        item.get("sup_email"), item.get("sup_phone"), "supervisor"
                    )
                    if sup:
                        GroupSupervisors.objects.create(group=group, user=sup, type="supervisor")

                created_count += 1

        return Response({"created_projects": created_count, "status": "success"})

    except Exception as e:
        return Response({"detail": f"خطأ في الصف {created_count + 3}: {str(e)}"}, status=400)